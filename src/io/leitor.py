"""
src/leitor.py
=============
Toda leitura de Excel. Zero lógica de negócio aqui.
"""

import os
from datetime import date, datetime
from typing import Optional

import openpyxl
import pandas as pd

from src.core.regras import ParamMes, Funcionario, MESES_PT


# ─── Batidas ─────────────────────────────────────────────────────────────────

def ler_batidas(caminho: str) -> pd.DataFrame:
    """
    Lê batidas.xlsx (todas as abas que tenham as colunas obrigatórias).
    Retorna DataFrame com: NOME, CPF, EMPRESA, TIPO_EVENTO, DATA_HORA, DATA
    """
    xl     = pd.ExcelFile(caminho)
    frames = []
    for aba in xl.sheet_names:
        try:
            df = pd.read_excel(caminho, sheet_name=aba)
            if {'NOME', 'TIPO_EVENTO', 'DATA_HORA'}.issubset(df.columns):
                frames.append(df)
        except Exception:
            pass

    if not frames:
        raise ValueError('Nenhuma aba válida em batidas.xlsx')

    df = pd.concat(frames, ignore_index=True)
    df['DATA_HORA'] = pd.to_datetime(df['DATA_HORA'])
    df['DATA']      = df['DATA_HORA'].dt.date
    df['NOME']      = df['NOME'].str.strip()
    return df


# ─── Parâmetros do mês ───────────────────────────────────────────────────────

def ler_params_mes(wb: openpyxl.Workbook, mes_num: int) -> ParamMes:
    """Lê PAR/ÍMPAR da aba MÊS para o mês informado."""
    ws = wb['MÊS']
    for row in ws.iter_rows(min_row=2, max_row=14, values_only=True):
        nome = str(row[0] or '').upper().strip()
        if MESES_PT.get(nome) == mes_num:
            return ParamMes(
                numero   = mes_num,
                nome     = nome,
                max_par  = int(row[7]),
                max_impar= int(row[8]),
            )
    raise ValueError(f'Mês {mes_num} não encontrado na aba MÊS')


# ─── Admissões parciais ──────────────────────────────────────────────────────

def ler_admissoes(wb: openpyxl.Workbook) -> dict[str, date]:
    """
    Lê aba ADMISSÕES.
    Retorna { nome_norm → data_inicio }
    """
    ws = wb['ADMISSÕES']
    resultado = {}
    for row in ws.iter_rows(min_row=3, max_row=200, values_only=True):
        nome  = row[2]
        inicio = row[4]
        if nome and inicio:
            nome_norm = str(nome).strip().lower()
            if isinstance(inicio, datetime):
                resultado[nome_norm] = inicio.date()
            elif isinstance(inicio, date):
                resultado[nome_norm] = inicio
    return resultado


# ─── Mapa de funcionários no FATURAMENTO ─────────────────────────────────────

def ler_funcionarios_faturamento(wb: openpyxl.Workbook,
                                  admissoes: dict[str, date]) -> list[Funcionario]:
    """
    Lê aba FATURAMENTO e retorna lista de Funcionario.
    Colunas: D=Nome(4), C=Matrícula(3), L=Escala(12)
    """
    ws    = wb['FATURAMENTO']
    funcs = []

    for row in ws.iter_rows(min_row=5, max_row=667):
        nome_cell   = row[3]   # coluna D
        matr_cell   = row[2]   # coluna C
        escala_cell = row[11]  # coluna L

        if not nome_cell.value:
            continue

        nome      = str(nome_cell.value).strip()
        nome_norm = nome.lower()
        matricula = matr_cell.value if matr_cell.value else None

        escala = str(escala_cell.value or '').upper()
        grupo  = 'ÍMPAR' if ('ÍMPAR' in escala or 'IMPAR' in escala) else 'PAR'

        admissao = admissoes.get(nome_norm)
        # Também tenta pela coluna J (data admissão) como fallback
        if admissao is None:
            adm_cell = row[9]  # coluna J
            if adm_cell.value:
                v = adm_cell.value
                admissao = v.date() if isinstance(v, datetime) else v

        funcs.append(Funcionario(
            nome       = nome,
            nome_norm  = nome_norm,
            matricula  = matricula,
            grupo      = grupo,
            admissao   = admissao,
            linha_excel= nome_cell.row,
        ))

    return funcs


# ─── Ausências registradas no ponto ──────────────────────────────────────────

def extrair_ausencias(df: pd.DataFrame) -> dict[str, list[dict]]:
    """
    Extrai eventos FALTA/ATESTAD por funcionário.
    Retorna { nome_norm → [{'data': date, 'motivo': str}] }
    """
    eventos_ausencia = df[df['TIPO_EVENTO'].isin(['FALTA', 'ATESTAD'])].copy()
    resultado = {}

    for _, row in eventos_ausencia.iterrows():
        nn     = row['NOME'].lower()
        motivo = 'Atestado Médico' if row['TIPO_EVENTO'] == 'ATESTAD' else 'Falta Injustificada'
        resultado.setdefault(nn, []).append({'data': row['DATA'], 'motivo': motivo})

    for nn in resultado:
        resultado[nn].sort(key=lambda x: x['data'])

    return resultado


# ─── Validação de arquivos de entrada ────────────────────────────────────────

def validar_inputs(pasta: str) -> dict:
    """
    Verifica se batidas.xlsx e o faturamento estão na pasta.
    Retorna { ok, batidas, faturamento, erros }
    """
    res = {'ok': False, 'batidas': None, 'faturamento': None, 'erros': []}

    if not os.path.exists(pasta):
        res['erros'].append(f'Pasta não encontrada: {pasta}')
        return res

    arquivos = os.listdir(pasta)

    bat = [f for f in arquivos if 'batidas' in f.lower() and f.endswith('.xlsx')]
    if bat:
        res['batidas'] = os.path.join(pasta, bat[0])
    else:
        res['erros'].append('batidas.xlsx não encontrado')

    fat = [f for f in arquivos
           if 'faturamento' in f.lower() and 'posto' in f.lower()
           and f.endswith('.xlsx') and 'processado' not in f.lower()]
    if fat:
        res['faturamento'] = os.path.join(pasta, fat[0])
    else:
        res['erros'].append('Faturamento_posto_de_trabalho_*.xlsx não encontrado')

    res['ok'] = not res['erros']
    return res

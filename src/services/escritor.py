"""
src/escritor.py
===============
Toda escrita em Excel. Zero lógica de negócio aqui.
"""

from datetime import datetime, date
from typing import Optional

import openpyxl

from src.core.regras import ResultadoFuncionario


# Colunas da aba FATURAMENTO
COL_INDISP_INI = 16   # P — Indisponibilidade Inicial
COL_INDISP_FIM = 17   # Q — Indisponibilidade Final
COL_TOT_INDISP = 19   # S — Total Indisponibilidade
COL_TOT_DISP   = 20   # T — Total Disponibilidade

# Colunas da aba INDISPONIBILIDADE
COL_IND_MATR    = 2   # B
COL_IND_NOME    = 3   # C
COL_IND_MOTIVO  = 4   # D
COL_IND_INI     = 5   # E
COL_IND_FIM     = 6   # F
COL_IND_TOTAL   = 7   # G
COL_IND_RETORNO = 8   # H


def _dt(d) -> Optional[datetime]:
    if d is None:
        return None
    if isinstance(d, datetime):
        return d
    if isinstance(d, date):
        return datetime(d.year, d.month, d.day)
    return None


def gravar_resultado_faturamento(ws, resultado: ResultadoFuncionario):
    """Grava disponibilidade e indisponibilidade na linha do funcionário."""
    linha = resultado.func.linha_excel

    ws.cell(row=linha, column=COL_TOT_DISP).value = resultado.disponibilidade

    if resultado.indisponibilidade > 0:
        ws.cell(row=linha, column=COL_TOT_INDISP).value = resultado.indisponibilidade
        aus = resultado.ausencias
        ini = _dt(aus[0]['data'])  if aus else None
        fim = _dt(aus[-1]['data']) if aus else None
        ws.cell(row=linha, column=COL_INDISP_INI).value = ini
        ws.cell(row=linha, column=COL_INDISP_FIM).value = fim
    else:
        # Limpa resíduos de meses anteriores
        ws.cell(row=linha, column=COL_TOT_INDISP).value = None
        ws.cell(row=linha, column=COL_INDISP_INI).value = None
        ws.cell(row=linha, column=COL_INDISP_FIM).value = None


def preencher_aba_indisponibilidade(wb: openpyxl.Workbook,
                                    resultados: list[ResultadoFuncionario]):
    """
    Limpa e repreenche a aba INDISPONIBILIDADE com os dados calculados.
    Agrupa ausências consecutivas (até 2 dias de gap) em um período.
    """
    ws = wb['INDISPONIBILIDADE']

    # Limpa a partir da linha 3
    for row in ws.iter_rows(min_row=3, max_row=300):
        for cell in row:
            cell.value = None

    linha = 3
    for res in resultados:
        if not res.ausencias:
            continue

        for periodo in _agrupar_periodos(res.ausencias):
            ws.cell(row=linha, column=COL_IND_MATR).value    = res.func.matricula
            ws.cell(row=linha, column=COL_IND_NOME).value    = res.func.nome
            ws.cell(row=linha, column=COL_IND_MOTIVO).value  = periodo['motivo']
            ws.cell(row=linha, column=COL_IND_INI).value     = _dt(periodo['inicio'])
            ws.cell(row=linha, column=COL_IND_FIM).value     = _dt(periodo['fim'])
            ws.cell(row=linha, column=COL_IND_TOTAL).value   = periodo['total']
            ws.cell(row=linha, column=COL_IND_RETORNO).value = _dt(periodo['retorno'])
            linha += 1


def _agrupar_periodos(ausencias: list[dict]) -> list[dict]:
    """
    Agrupa ausências consecutivas (gap ≤ 2 dias) em períodos.
    Retorna lista de { inicio, fim, total, motivo, retorno }
    """
    from datetime import timedelta

    if not ausencias:
        return []

    periodos = []
    ini   = ausencias[0]['data']
    fim   = ausencias[0]['data']
    mots  = [ausencias[0]['motivo']]

    for item in ausencias[1:]:
        gap = (item['data'] - fim).days
        if gap <= 2:
            fim = item['data']
            mots.append(item['motivo'])
        else:
            periodos.append(_montar_periodo(ini, fim, mots))
            ini, fim, mots = item['data'], item['data'], [item['motivo']]

    periodos.append(_montar_periodo(ini, fim, mots))
    return periodos


def _montar_periodo(ini, fim, mots: list[str]) -> dict:
    from datetime import timedelta
    motivo  = 'Atestado Médico' if 'Atestado Médico' in mots else 'Falta Injustificada'
    retorno = fim + timedelta(days=2)
    return {'inicio': ini, 'fim': fim, 'total': len(mots),
            'motivo': motivo, 'retorno': retorno}

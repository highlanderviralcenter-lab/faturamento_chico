import os
import sys
import shutil
import math
from datetime import date, datetime, timedelta

import openpyxl
import pandas as pd


PASTA = os.path.dirname(os.path.abspath(__file__))

MESES_PT = {
    'JANEIRO':1,'FEVEREIRO':2,'MARÇO':3,'ABRIL':4,'MAIO':5,'JUNHO':6,
    'JULHO':7,'AGOSTO':8,'SETEMBRO':9,'OUTUBRO':10,'NOVEMBRO':11,'DEZEMBRO':12,
}

COL_INDISP_INI = 16
COL_INDISP_FIM = 17
COL_TOT_INDISP = 19
COL_TOT_DISP   = 20


def achar_arquivos():
    arquivos = os.listdir(PASTA)
    batidas  = [f for f in arquivos if 'batidas' in f.lower() and f.endswith('.xlsx')]
    fat      = [f for f in arquivos
                if 'faturamento' in f.lower() and 'posto' in f.lower()
                and f.endswith('.xlsx') and 'processado' not in f.lower()]
    if not batidas:
        print('nao achei batidas.xlsx aqui')
        sys.exit(1)
    if not fat:
        print('nao achei Faturamento_posto_*.xlsx aqui')
        sys.exit(1)
    return os.path.join(PASTA, batidas[0]), os.path.join(PASTA, fat[0])


def ler_batidas(caminho):
    df = pd.read_excel(caminho)
    df['DATA_HORA'] = pd.to_datetime(df['DATA_HORA'])
    df['DATA']      = df['DATA_HORA'].dt.date
    df['NOME']      = df['NOME'].str.strip()
    return df


def ler_params_mes(wb, mes_num):
    ws = wb['MÊS']
    for row in ws.iter_rows(min_row=2, max_row=14, values_only=True):
        nome = str(row[0] or '').upper().strip()
        if MESES_PT.get(nome) == mes_num:
            return {'numero': mes_num, 'nome': nome,
                    'max_par': int(row[7]), 'max_impar': int(row[8]),
                    'total': int(row[7]) + int(row[8])}
    raise ValueError(f'mes {mes_num} nao ta na aba MES')


def ler_admissoes(wb):
    ws  = wb['ADMISSÕES']
    res = {}
    for row in ws.iter_rows(min_row=3, max_row=200, values_only=True):
        nome = row[2]
        dt   = row[4]
        if nome and dt:
            d = dt.date() if isinstance(dt, datetime) else dt
            res[str(nome).strip().lower()] = d
    return res


def ler_funcionarios(wb, admissoes):
    ws    = wb['FATURAMENTO']
    funcs = []
    for row in ws.iter_rows(min_row=5, max_row=667):
        nc = row[3]
        if not nc.value:
            continue
        nome      = str(nc.value).strip()
        nome_norm = nome.lower()
        matricula = row[2].value
        escala    = str(row[11].value or '').upper()
        grupo     = 'ÍMPAR' if ('ÍMPAR' in escala or 'IMPAR' in escala) else 'PAR'
        admissao  = admissoes.get(nome_norm)
        if admissao is None:
            v = row[9].value
            if v:
                admissao = v.date() if isinstance(v, datetime) else v
        funcs.append({'nome': nome, 'norm': nome_norm, 'matricula': matricula,
                      'grupo': grupo, 'admissao': admissao, 'linha': nc.row})
    return funcs


def extrair_ausencias(df):
    res = {}
    for _, row in df[df['TIPO_EVENTO'].isin(['FALTA', 'ATESTAD'])].iterrows():
        nn     = row['NOME'].lower()
        motivo = 'Atestado Medico' if row['TIPO_EVENTO'] == 'ATESTAD' else 'Falta'
        res.setdefault(nn, []).append({'data': row['DATA'], 'motivo': motivo})
    for nn in res:
        res[nn].sort(key=lambda x: x['data'])
    return res


def auditar_jornadas(df):
    registros = []
    for (nome, data), g in df.groupby(['NOME', 'DATA']):
        login_row  = g[g['TIPO_EVENTO'] == 'LOGIN']
        logout_row = g[g['TIPO_EVENTO'] == 'LOGOUT']
        if login_row.empty or logout_row.empty:
            continue
        login  = login_row['DATA_HORA'].iloc[0]
        logout = logout_row['DATA_HORA'].iloc[-1]
        # turno noturno: logout no dia seguinte
        if logout < login:
            logout = logout + timedelta(days=1)
        horas = (logout - login).total_seconds() / 3600
        registros.append({'nome': nome, 'data': data,
                          'login': login, 'logout': logout,
                          'horas': round(horas, 2), 'ok': horas >= 12.0})
    return registros


def max_admissao_parcial(max_grupo, admissao, params):
    inicio = date(2026, params['numero'], 1)
    if admissao <= inicio:
        return max_grupo
    dias = params['total'] - admissao.day + 1
    return math.ceil(max_grupo * dias / params['total'])


def calcular(logins, max_dias, parcial):
    if parcial and logins >= max_dias:
        return logins, 0
    return logins, max(max_dias - logins, 0)


def to_dt(d):
    if d is None:
        return None
    if isinstance(d, datetime):
        return d
    return datetime(d.year, d.month, d.day)


def gravar(ws, func, disponib, indisp, ausencias):
    linha = func['linha']
    ws.cell(row=linha, column=COL_TOT_DISP).value = disponib
    if indisp > 0:
        ws.cell(row=linha, column=COL_TOT_INDISP).value = indisp
        ws.cell(row=linha, column=COL_INDISP_INI).value = to_dt(ausencias[0]['data'])  if ausencias else None
        ws.cell(row=linha, column=COL_INDISP_FIM).value = to_dt(ausencias[-1]['data']) if ausencias else None
    else:
        ws.cell(row=linha, column=COL_TOT_INDISP).value = None
        ws.cell(row=linha, column=COL_INDISP_INI).value = None
        ws.cell(row=linha, column=COL_INDISP_FIM).value = None


def agrupar_periodos(ausencias):
    if not ausencias:
        return []
    periodos = []
    ini  = ausencias[0]['data']
    fim  = ausencias[0]['data']
    mots = [ausencias[0]['motivo']]
    for item in ausencias[1:]:
        if (item['data'] - fim).days <= 2:
            fim = item['data']
            mots.append(item['motivo'])
        else:
            periodos.append({'inicio': ini, 'fim': fim, 'total': len(mots),
                             'motivo': 'Atestado Medico' if 'Atestado Medico' in mots else 'Falta',
                             'retorno': fim + timedelta(days=2)})
            ini, fim, mots = item['data'], item['data'], [item['motivo']]
    periodos.append({'inicio': ini, 'fim': fim, 'total': len(mots),
                     'motivo': 'Atestado Medico' if 'Atestado Medico' in mots else 'Falta',
                     'retorno': fim + timedelta(days=2)})
    return periodos


def preencher_indisponibilidade(wb, resultados):
    ws = wb['INDISPONIBILIDADE']
    for row in ws.iter_rows(min_row=3, max_row=300):
        for cell in row:
            cell.value = None
    linha = 3
    for r in resultados:
        if not r['ausencias']:
            continue
        for p in agrupar_periodos(r['ausencias']):
            ws.cell(row=linha, column=2).value = r['matricula']
            ws.cell(row=linha, column=3).value = r['nome']
            ws.cell(row=linha, column=4).value = p['motivo']
            ws.cell(row=linha, column=5).value = to_dt(p['inicio'])
            ws.cell(row=linha, column=6).value = to_dt(p['fim'])
            ws.cell(row=linha, column=7).value = p['total']
            ws.cell(row=linha, column=8).value = to_dt(p['retorno'])
            linha += 1


def main():
    bat_path, fat_path = achar_arquivos()

    print(f'batidas:     {os.path.basename(bat_path)}')
    print(f'faturamento: {os.path.basename(fat_path)}')
    print()

    df_bat  = ler_batidas(bat_path)
    mes_num = int(pd.to_datetime(df_bat['DATA_HORA']).dt.month.mode()[0])

    # ── auditoria 12h ─────────────────────────────────────────────────────────
    jornadas = auditar_jornadas(df_bat)
    abaixo   = [j for j in jornadas if not j['ok']]

    if abaixo:
        print(f'jornadas abaixo de 12h  ({len(abaixo)} ocorrencias):')
        print(f'  {"nome":<35} {"data":<12} {"login":<7} {"logout":<7} horas')
        print('  ' + '-' * 65)
        for j in abaixo:
            print(f'  {j["nome"][:34]:<35} {str(j["data"]):<12} '
                  f'{j["login"].strftime("%H:%M"):<7} '
                  f'{j["logout"].strftime("%H:%M"):<7} '
                  f'{int(j["horas"])}h{int((j["horas"] % 1) * 60):02d}min')
        print()
    else:
        print('todas as jornadas com 12h ou mais')
        print()

    # ── faturamento ───────────────────────────────────────────────────────────
    ts    = datetime.now().strftime('%m_%Y_%H%M')
    saida = os.path.join(PASTA, f'Faturamento_PROCESSADO_{ts}.xlsx')
    shutil.copy2(fat_path, saida)

    wb_ro = openpyxl.load_workbook(saida, data_only=True)
    wb    = openpyxl.load_workbook(saida)

    params    = ler_params_mes(wb_ro, mes_num)
    admissoes = ler_admissoes(wb_ro)
    funcs     = ler_funcionarios(wb_ro, admissoes)
    ausencias = extrair_ausencias(df_bat)

    diff     = params['max_impar'] - params['max_par']
    esperado = 1 if params['total'] % 2 == 1 else 0
    if diff != esperado:
        print(f"ERRO: PAR={params['max_par']} IMPAR={params['max_impar']} nao fecha pra {params['total']} dias")
        sys.exit(1)

    print(f"{params['nome']}  {params['total']} dias  PAR={params['max_par']}  IMPAR={params['max_impar']}")
    print()

    logins_map = (
        df_bat[df_bat['TIPO_EVENTO'] == 'LOGIN']
        .groupby(df_bat['NOME'].str.lower().str.strip())['DATA']
        .nunique()
        .to_dict()
    )

    max_grupo  = {'PAR': params['max_par'], 'ÍMPAR': params['max_impar']}
    ws_fat     = wb['FATURAMENTO']
    resultados = []
    avisos     = []

    print(f'{"nome":<38} {"grupo":<6} {"logins":>6}  {"disp":>5}  {"indisp":>6}')
    print('-' * 68)

    for f in funcs:
        logins = logins_map.get(f['norm'])
        if logins is None:
            if f['norm'] not in ('legendas', 'contratado'):
                avisos.append(f"sem batidas: {f['nome']}")
            continue

        max_std  = max_grupo[f['grupo']]
        inicio   = date(2026, mes_num, 1)
        parcial  = f['admissao'] is not None and f['admissao'] > inicio
        max_dias = max_admissao_parcial(max_std, f['admissao'], params) if parcial else max_std

        if not parcial and logins > max_std:
            avisos.append(f"grupo errado? {f['nome']}: {logins} logins > max {max_std}")

        disp, indisp = calcular(logins, max_dias, parcial)
        aus = ausencias.get(f['norm'], [])

        if indisp > 0 and indisp > len(aus):
            avisos.append(f"sem motivo completo: {f['nome']} — {indisp} dias, {len(aus)} registro(s)")

        flag = '  *' if indisp > 0 else ''
        print(f"{f['nome'][:37]:<38} {f['grupo']:<6} {logins:>6}  {disp:>5}  {indisp:>6}{flag}")

        gravar(ws_fat, f, disp, indisp, aus)
        resultados.append({**f, 'disp': disp, 'indisp': indisp, 'ausencias': aus})

    preencher_indisponibilidade(wb, resultados)
    wb.save(saida)

    print()
    if avisos:
        print('avisos:')
        for a in avisos:
            print(f'  {a}')
        print()

    print(f'salvo: {os.path.basename(saida)}')


if __name__ == '__main__':
    main()